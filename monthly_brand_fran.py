import requests
import json
import pandas as pd
import datetime as dt
from ordered_set import OrderedSet
from topic_senti import sentiOpenTopic
import os
import warnings
warnings.filterwarnings("ignore")
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
headers = {
    'Cookie': 'language=en_US; gempollId=dc1075a5db4b4f56b49570fb5bda2d81',
    'Host': 'sl.cn-siit.com',
    'Origin': 'https://sl.cn-siit.com',
    'Referer': 'https://sl.cn-siit.com/overview?project=52&dashboard=66&from=2019-10-1&to=2019-10-31&selectPlatform=1&query=120,122,124,126,128,658,666,674,680,682,696,698',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.70 Safari/537.36'
}

id_query ={
            '66': '120,122,124,126,128,658,666,674,680,682,696,698',#brand
           '102': '130,526,528,538,542,544,654,668,684,686,692,700',
           '104': '260,278,292,520,534,540,546,548,552,558,676,694',
           '106': '284,286,522,530,536,550,554,556,560,562,564,690',
           # franchise
           '68': '132,134,140,234,236,238,240,242,244,246,338,466',
           '108': '154,200,202,204,206,214,216,218,222,248,254,256',
           '110': '152,208,210,212,224,226,228,230,232,250,252,362',
           '112': '146,198,220,312,314,332,358,360,384,452,458,566',
            # campaign
           # '72': '1020,1022',
           #  '168': '1144,1145'
    '72': '1142,1022,1144,1492,1494',
    # '168': '1492,1494'
    }

id_type = {'66': 'brand',
           '102': 'brand',
           '104': 'brand',
           '106': 'brand',
           '68': 'franchise',
           '108': 'franchise',
           '110': 'franchise',
           '112': 'franchise',
           '72': 'campaign',
           '168': 'campaign'}

# form_data示例
form_data = {
    'dashboardType': '0',
    'id': '68',#68,152, 66
    'type': '2',
    'beginTime': '2019-10-20',
    #可变日期
    'endTime': '2019-10-26',
    'query': '132,134,140,234,236,238,240,242,244,246,338,466'}

#写入excel指定位置，输入的start_row最小值为2，start_nol最小值为1
def iloc_write(st_name, df, start_row = 2, start_nol = 1):
    #写入主体内容
    for rcol in range(start_row, df.shape[0]+start_row):
        for ncol in range(start_nol, df.shape[1]+start_nol):
            col_en = get_column_letter(ncol)
            # print(rcol, ncol)
            _ = st_name.cell(rcol, ncol, value='{}'.format(df.iloc[rcol-start_row, ncol-start_nol]))
            st_name['{}{}'.format(col_en, rcol)].font = Font(name=u'微软雅黑', size=10)
            try:
                #将数字转为int形式
                _.value = int(_.value)
                _.number_format = '#,##0'
            except:
                try:
                    _.value = float(_.value)
                    _.number_format = '0.00%'
                except:
                    pass

    #写入header
    for rcol in range(start_row-1, start_row):
        for ncol in range(start_nol, len(df.columns)+start_nol):
            col_en = get_column_letter(ncol)
            _ = st_name.cell(rcol, ncol, value='{}'.format(df.columns[ncol-start_nol]))
            st_name['{}{}'.format(col_en, rcol)].font = Font(name=u'微软雅黑', size=10)




def daily_plt(headers, id, query, dayslist, type):
    url ='https://sl.cn-siit.com/postDetail'
    form_data = {
        'dashboardType': '0',
        'id': str(id),
        'type': '2',
        'beginTime': dayslist[0],
        # 可变日期
        'endTime': dayslist[-1],
        'query': str(query)
    }
    temp = pd.DataFrame()
    rq = requests.post(url,data=form_data,headers=headers)
    cao = rq.content.decode()
    cnm = json.loads(cao)['page']['platform_list']
    for i in cnm:
        # 取出每天的平台声量
        for adate in i['date_list']:
            c = {'date':[adate],'platform':[i['name']],'volume':i[adate], 'query':form_data['query'], 'type':type}
            temp = temp.append(pd.DataFrame(c))
    return temp


# ec, weibo, wechat, news, bbs, vedio, qa
platform_id = ['5', '7', '6', '3', '1', '4', '2']
# ec/total情感值
def sentiOpen(headers, id, query, dayslist, type, plt_id):
    temp = pd.DataFrame()
    url = 'https://sl.cn-siit.com/postSentimentOpen'
    if plt_id in platform_id:
        form_data = {
            'dashboardType': '0',
            'id': id,
            'type': '2',
            'beginTime': dayslist[0],
            # 可变日期
            'endTime': dayslist[-1],
            'platformId': plt_id,
            'query': query
        }
        plt = 'ec'
    elif plt_id == '7,6,3,1,4,2':
        form_data = {
            'dashboardType': '0',
            'id': id,
            'type': '2',
            'beginTime': dayslist[0],
            # 可变日期
            'endTime': dayslist[-1],
            'platformId': plt_id,
            'query': query
        }
        plt = 'social'
    else:
        form_data = {
            'dashboardType': '0',
            'id': id,
            'type': '2',
            'beginTime': dayslist[0],
            # 可变日期
            'endTime': dayslist[-1],
            'query': query
        }
        plt = 'all'


    rq = requests.post(url, data=form_data, headers=headers)
    cao = rq.content.decode()
    cnm = json.loads(cao)['list']
    for i in cnm:
        c = {'date': i['date'], '正面': i['total1'], '负面': i['total3'], '中性': i['total2'],
             'query':[form_data['query']], 'type':type, 'plt': plt}
        temp = temp.append(pd.DataFrame(c))
    return temp

# brand首页表现
def topic_buzz(headers, id, query, dayslist, type, pltid = '5,7,6,3,1,4,2'):
    df = pd.DataFrame()
    url = r'https://sl.cn-siit.com/postOverviewBrand'
    form_data = {
        'dashboardType': '0',
        'id': str(id),
        'type': '2',
        'beginTime': dayslist[0],
        # 可变日期
        'endTime': dayslist[-1],
        'query': str(query),
        'platformId': str(pltid)
    }

    rq = requests.post(url, data=form_data, headers=headers)
    cao = rq.content.decode()
    cnm = json.loads(cao)
    for i in cnm:
        plat = i['platform_list']
        topic_list = i['topic_list']
        temp = {}
        temp['name'] = i['brand']
        temp['声量'] = i['total_buzz']['total']
        temp['声量趋势'] = i['total_buzz']['percent']
        temp['情感值'] = i['nsr']['like']
        temp['情感值趋势'] = i['nsr']['percent']
        num = [x for x in range(len(plat))]
        for nc in num:
            temp[plat[nc]['platform']] = plat[nc]['total']
            temp['%s占比' % plat[nc]['platform']] = plat[nc]['proportion']
        # temp[plat[0]['platform']] = plat[0]['total']
        # temp['%s占比' % plat[0]['platform']] = plat[0]['proportion']
        # temp[plat[1]['platform']] = plat[1]['total']
        # temp['%s占比' % plat[1]['platform']] = plat[1]['proportion']
        # temp[plat[2]['platform']] = plat[2]['total']
        # temp['%s占比' % plat[2]['platform']] = plat[2]['proportion']
        # temp[plat[3]['platform']] = plat[3]['total']
        # temp['%s占比' % plat[3]['platform']] = plat[3]['proportion']
        # temp[plat[4]['platform']] = plat[4]['total']
        # temp['%s占比' % plat[4]['platform']] = plat[4]['proportion']
        # temp[plat[5]['platform']] = plat[5]['total']
        # temp['%s占比' % plat[5]['platform']] = plat[5]['proportion']
        # temp[plat[6]['platform']] = plat[6]['total']
        # temp['%s占比' % plat[6]['platform']] = plat[6]['proportion']
        temp['互动量'] = i['engagement']['total']
        temp['互动量趋势'] = i['engagement']['percent']
        temp[topic_list[0]['topic']] = topic_list[0]['total']
        temp['%s趋势' % topic_list[0]['topic']] = topic_list[0]['percent']
        temp[topic_list[1]['topic']] = topic_list[1]['total']
        temp['%s趋势' % topic_list[1]['topic']] = topic_list[1]['percent']
        temp[topic_list[2]['topic']] = topic_list[2]['total']
        temp['%s趋势' % topic_list[2]['topic']] = topic_list[2]['percent']
        temp[topic_list[3]['topic']] = topic_list[3]['total']
        temp['%s趋势' % topic_list[3]['topic']] = topic_list[3]['percent']
        temp[topic_list[4]['topic']] = topic_list[4]['total']
        temp['%s趋势' % topic_list[4]['topic']] = topic_list[4]['percent']
        temp['type'] = [type]
        temp['query'] = query
        df = df.append(pd.DataFrame(temp))
    # print(df.columns)
    return df

def DetailTopic(headers, id, query, dayslist, pltid ='5,7,6,3,1,4,2'):
    url = 'https://sl.cn-siit.com/postDetail'
    form_data = {
        'dashboardType': '0',
        'id': str(id),
        'type': '2',
        'beginTime': dayslist[0],
        'platformId': str(pltid),
        # 可变日期
        'endTime': dayslist[-1],
        'query': str(query)
    }
    rq = requests.post(url, data=form_data, headers=headers)
    # print(form_data)
    cao = rq.content.decode()
    cnm = json.loads(cao)['page']['topic_a_data']['topic_list']
    df = pd.DataFrame()
    # print(cnm)
    for i in cnm:
        first_name = i['topic']
        first_num = i['total']
        # 2级属性
        second_plat = i['topic_sub_list']
        for s in second_plat:
            second_name = s['topic']
            second_num = s['total']
            #3级属性
            third_plat = s['topic_sub_list']
            for t in third_plat:
                rnm = {}
                third_name = t['topic']
                third_num = t['total']
                rnm['1级属性'] = [first_name]
                rnm['1级属性总量'] = first_num
                rnm['2级属性'] = second_name
                rnm['2级属性总量'] = second_num
                rnm['3级属性'] = third_name
                rnm['3级属性总量'] = third_num
                rnm['query'] = query
                rnm['pltid'] = str(pltid)
                # print(rnm)

                df = df.append(pd.DataFrame(rnm))
    return df

def DetailTopicFran(headers, id, query, dayslist, pltid ='5,7,6,3,1,4,2'):
    url = 'https://sl.cn-siit.com/postDetail'
    form_data = {
        'dashboardType': '0',
        'id': str(id),
        'type': '2',
        'beginTime': dayslist[0],
        'platformId': str(pltid),
        # 可变日期
        'endTime': dayslist[-1],
        'query': str(query)
    }
    rq = requests.post(url, data=form_data, headers=headers)
    # print(form_data)
    cao = rq.content.decode()
    cnm = json.loads(cao)['page']['topic_a_data']['topic_list']
    df = pd.DataFrame()
    # print(cnm)
    for i in cnm:
        first_name = i['topic']
        first_num = i['total']
        # 2级属性
        second_plat = i['topic_sub_list']
        for s in second_plat:
            second_name = s['topic']
            second_num = s['total']
            #3级属性
            rnm = {}
            rnm['1级属性'] = [first_name]
            rnm['1级属性总量'] = first_num
            rnm['2级属性'] = second_name
            rnm['2级属性总量'] = second_num
            rnm['query'] = query
            rnm['pltid'] = str(pltid)
            df = df.append(pd.DataFrame(rnm))
    return df

# franchise 首页表现
def get_fran_topic(headers, id, query, dayslist, type, pltid = '5,7,6,3,1,4,2'):
    df = pd.DataFrame()
    url = 'https://sl.cn-siit.com/postOverviewFranchise'
    form_data = {
        'dashboardType': '0',
        'id': str(id),
        'type': '3',
        'beginTime': dayslist[0],
        # 可变日期
        'endTime': dayslist[-1],
        'query': str(query),
        'platformId': str(pltid)
    }
    rq = requests.post(url, data=form_data, headers=headers)
    cao = rq.content.decode()
    cnm = json.loads(cao)
    for i in cnm:
        plat = i['platform_list']
        topic_list = i['topic_list']
        temp = {}
        temp['name'] = [i['franchise']]
        temp['声量'] = i['buzz']['total']
        temp['声量趋势'] = i['buzz']['percent']

        temp['情感值'] = i['nsr']['like']
        temp['情感值趋势'] = i['nsr']['percent']

        num = [x for x in range(len(plat))]

        for nc in num:
            temp[plat[nc]['platform']] = plat[nc]['total']
            temp['%s占比' % plat[nc]['platform']] = plat[nc]['proportion']
            # temp[plat[0]['platform']] = plat[0]['total']
            # temp['%s占比' % plat[0]['platform']] = plat[0]['proportion']
            # temp[plat[1]['platform']] = plat[1]['total']
            # temp['%s占比' % plat[1]['platform']] = plat[1]['proportion']
            # temp[plat[2]['platform']] = plat[2]['total']
            # temp['%s占比' % plat[2]['platform']] = plat[2]['proportion']
            # temp[plat[3]['platform']] = plat[3]['total']
            # temp['%s占比' % plat[3]['platform']] = plat[3]['proportion']
            # temp[plat[4]['platform']] = plat[4]['total']
            # temp['%s占比' % plat[4]['platform']] = plat[4]['proportion']
            # temp[plat[5]['platform']] = plat[5]['total']
            # temp['%s占比' % plat[5]['platform']] = plat[5]['proportion']
            # temp[plat[6]['platform']] = plat[6]['total']
            # temp['%s占比' % plat[6]['platform']] = plat[6]['proportion']
        temp['type'] = type
        for i in range(9):
            temp[topic_list[i]['topic']] = topic_list[i]['total']
            temp['%s趋势' % topic_list[i]['topic']] = topic_list[i]['percent']
        df = df.append(pd.DataFrame(temp))
    return df

# 获取到天的声量
def daily_buzz(headers, id, query, dayslist, type):
    df = pd.DataFrame()
    url = 'https://sl.cn-siit.com/postOverviewUp'
    form_data = {
        'dashboardType': '0',
        'id': str(id),
        'type': '2',
        'beginTime': dayslist[0],
        # 可变日期
        'endTime': dayslist[-1],
        'query': str(query)}
    rq = requests.post(url, data=form_data, headers=headers)
    rqCont = rq.content.decode()
    # print(rqCont)
    contDict = json.loads(rqCont)['middle_list']
    for i in contDict:
        name = i['name']
        data_list = i['date_list']
        for t in range(len(data_list)):
            date = i['date_list'][t]
            value = i[date]
            c = {'name': [name], 'date': date, 'volume': value, 'query': form_data['query'], 'type': type}
            df = df.append(pd.DataFrame(c))
    return df

def topic_buzz_4(topicDf, sheet2):
    nike_topic = topicDf[topicDf['name'].isin(['Nike', 'Adidas', 'Lining', 'Anta'])]
    nike_topic = nike_topic[['name', 'Product', 'Purchase Intent', 'Brand Reputation', 'Campaign', 'Celebrity & KOL']]
    nike_topic['Total'] = list(
        map(lambda a, b, c, d, e: int(a) + int(b) + int(c) + int(d) + int(e), nike_topic['Product'],
            nike_topic['Purchase Intent'],
            nike_topic['Brand Reputation'], nike_topic['Campaign'], nike_topic['Celebrity & KOL']))
    for col in ['Product', 'Purchase Intent', 'Brand Reputation', 'Campaign', 'Celebrity & KOL']:
        nike_topic[col] = nike_topic[col].astype(int)
        nike_topic['{}_Pct'.format(col)] = list(map(lambda x, y: float(x/y), nike_topic[col], nike_topic['Total']))

    topic_Acols = OrderedSet(['name', 'Product', 'Purchase Intent', 'Brand Reputation', 'Campaign',
       'Celebrity & KOL', 'Total', 'Product_Pct', 'Purchase Intent_Pct',
       'Brand Reputation_Pct', 'Campaign_Pct', 'Celebrity & KOL_Pct'])
    topic_cols = list(topic_Acols & set(nike_topic.columns))
    nike_topic = nike_topic[topic_cols]

    iloc_write(sheet2, nike_topic, 3, 2)

def topic_buzz_4_detail(topicDf, sheet2):
    nike_topic = topicDf[topicDf['name'].isin(['Nike', 'Adidas', 'Lining', 'Anta'])]
    nike_topic = nike_topic[['name', 'Product', 'Purchase Intent', 'Brand', 'Campaign/Events', 'Category']]
    nike_topic['Total'] = list(
        map(lambda a, b, c, d, e: int(a) + int(b) + int(c) + int(d) + int(e), nike_topic['Product'], nike_topic['Purchase Intent'],
            nike_topic['Brand'], nike_topic['Campaign/Events'], nike_topic['Category']))
    for col in ['Product', 'Purchase Intent', 'Brand', 'Campaign/Events', 'Category']:
        nike_topic[col] = nike_topic[col].astype(int)
        nike_topic['{}_Pct'.format(col)] = list(map(lambda x, y: float(x/y), nike_topic[col], nike_topic['Total']))

    topic_Acols = OrderedSet(['name', 'Product', 'Purchase Intent', 'Brand', 'Campaign/Events',
       'Category', 'Total', 'Product_Pct', 'Purchase Intent_Pct',
       'Brand_Pct', 'Campaign/Events_Pct', 'Category_Pct'])
    topic_cols = list(topic_Acols & set(nike_topic.columns))
    nike_topic = nike_topic[topic_cols]

    iloc_write(sheet2, nike_topic, 3, 2)

# 整个时间段的情感计算
def sent_from_daily(senDf, writer):
    brandSent = senDf.pivot_table(index=['name', 'query', 'plt', 'type'],values=['正面', '负面', '中性'], aggfunc='sum').reset_index()
    brandSent['PSR'] = list(map(lambda x, y: x / (x + y) if (x + y) > 0 else '-', brandSent['正面'], brandSent['负面']))
    ec_sent = brandSent[brandSent['plt']=='ec']
    all_sent = brandSent[brandSent['plt']=='all']
    social_sent = brandSent[brandSent['plt']=='social']
    for col in ['正面', '负面', '中性', 'PSR']:
        if col != 'PSR':
            ec_sent[col] = ec_sent[col].astype(int)
            all_sent[col] = all_sent[col].astype(int)
            social_sent[col] = social_sent[col].astype(int)
        ec_sent = ec_sent.rename(columns={col: 'ec_{}'.format(col)})
        all_sent = all_sent.rename(columns={col: 'total_{}'.format(col)})
        social_sent = social_sent.rename(columns={col: 'social_{}'.format(col)})
    sent = all_sent.merge(ec_sent, on=['name', 'query', 'type'], how='left').merge(social_sent, on=['name', 'query', 'type'], how='left')
    sent = sent.fillna(0)
    # for col in ['正面', '负面', '中性']:
    #     sent['social_{}'.format(col)] = list(map(lambda x, y: int(x) - int(y), sent['total_{}'.format(col)], sent['ec_{}'.format(col)]))
    # sent['social_PSR'] = list(map(lambda x, y: x / (x + y) if (x + y) > 0 else '-', sent['social_正面'], sent['social_负面']))
    # 这里先把sent total写出
    sent = sent[['type', 'name', 'total_正面', 'total_中性', 'total_负面', 'total_PSR', 'social_正面', 'social_中性', 'social_负面', 'social_PSR',
                 'ec_正面', 'ec_中性', 'ec_负面', 'ec_PSR']]
    sent.to_excel(writer, sheet_name='总PSR', index=False)
    return sent


def fran_topic(franTopic):
    product_topic = franTopic[
        ['name', '声量', '电商', 'Sizing', 'Design', 'Price', 'Quality', 'Authenticity', 'Material', 'Flexibility',
         'Fitting', 'Comfort']]
    product_topic = product_topic.rename(columns={'声量': 'Buzz', '电商': 'EC'})
    product_topic = product_topic.fillna(0)
    product_topic['Social'] = list(map(lambda x, y: int(x) - int(y), product_topic['Buzz'], product_topic['EC']))
    # product_topic = product_topic.merge(sent[['name', 'total_PSR', 'social_PSR', 'ec_PSR']], on='name', how='left')
    # product_topic = product_topic[['name', 'Buzz', 'total_PSR', 'Social', 'social_PSR', 'EC', 'ec_PSR',
    #                                'Sizing', 'Design', 'Price', 'Quality', 'Authenticity', 'Material', 'Flexibility',
    #                                'Fitting', 'Comfort']]
    for col in ['Buzz', 'EC', 'Sizing', 'Design', 'Price', 'Quality', 'Authenticity', 'Material', 'Flexibility', 'Fitting', 'Comfort']:
        product_topic[col] = product_topic[col].fillna(0)
        product_topic[col] = product_topic[col].astype(int)
    product_topic['topic_sum'] = list(map(lambda a, b, c, d, e, f, g, h, i: int(a)+int(b)+int(c)+int(d)+int(e)+int(f)+int(g)+int(h)+int(i),
                                          product_topic['Sizing'], product_topic['Design'],product_topic['Price'],product_topic['Quality'],
                                          product_topic['Authenticity'],product_topic['Material'],product_topic['Flexibility'],product_topic['Fitting'],
                                          product_topic['Comfort']))
    for col in ['Sizing', 'Design', 'Price', 'Quality', 'Authenticity', 'Material', 'Flexibility', 'Fitting', 'Comfort']:
        product_topic['{}_Pct'.format(col)] = list(map(lambda x, y: float(x/y), product_topic[col], product_topic['topic_sum']))
    return product_topic

def fran_topic_detail(franTopic):
    product_topic = franTopic[
        ['name', 'Sizing', 'Design', 'Price', 'Quality', 'Authenticity', 'Material', 'Flexibility',
         'Fitting', 'Comfort']]
    product_topic = product_topic.fillna(0)
    for col in ['Sizing', 'Design', 'Price', 'Quality', 'Authenticity', 'Material', 'Flexibility', 'Fitting', 'Comfort']:
        product_topic[col] = product_topic[col].fillna(0)
        product_topic[col] = product_topic[col].astype(int)
    product_topic['topic_sum'] = list(map(lambda a, b, c, d, e, f, g, h, i: int(a)+int(b)+int(c)+int(d)+int(e)+int(f)+int(g)+int(h)+int(i),
                                          product_topic['Sizing'], product_topic['Design'],product_topic['Price'],product_topic['Quality'],
                                          product_topic['Authenticity'],product_topic['Material'],product_topic['Flexibility'],product_topic['Fitting'],
                                          product_topic['Comfort']))
    for col in ['Sizing', 'Design', 'Price', 'Quality', 'Authenticity', 'Material', 'Flexibility', 'Fitting', 'Comfort']:
        product_topic['{}_Pct'.format(col)] = list(map(lambda x, y: float(x/y) if y > 0 else 'nan', product_topic[col], product_topic['topic_sum']))
    return product_topic


# 输出nike, adidas, lining, anta四个品牌的到天声量
def daily_volume_from_plt(pltDf, sheet3):
    dailyPlt = pltDf[pltDf['name'].isin(['Nike', 'Lining', 'Adidas', 'Anta'])]
    # dailyPlt = df[df['name'].isin(['Nike', 'Lining', 'Adidas', 'Anta'])]
    dailyPlt['is_ec'] = dailyPlt['platform'].map(lambda x: 'EC' if str(x)=='电商' else 'SOCIAL')
    dailyPlt_copy = dailyPlt.copy()
    dailyPlt_copy['is_ec'] = 'TOTAL'
    dailyPlt = dailyPlt.append(dailyPlt_copy)
    dailyPlt['name_ec'] = list(map(lambda x, y: str(x) +'_'+ str(y), dailyPlt['is_ec'], dailyPlt['name']))
    dailyout = dailyPlt.pivot_table(index='date', columns='name_ec', values='volume', aggfunc='sum').reset_index()
    dailyout = dailyout[['date', 'TOTAL_Nike', 'TOTAL_Adidas', 'TOTAL_Lining', 'TOTAL_Anta',
                         'SOCIAL_Nike', 'SOCIAL_Adidas', 'SOCIAL_Lining', 'SOCIAL_Anta',
                         'EC_Nike', 'EC_Adidas', 'EC_Lining', 'EC_Anta']]
    # print(dailyout.head())
    iloc_write(sheet3, dailyout, 3, 2)

def sheet10_writer(product_topic, prOther, sheet10, detail_tag=False):
    if detail_tag:
        prTopic = product_topic.merge(prOther[['name', 'total', 'total_PSR', 'social', 'social_PSR', '电商', 'ec_PSR']],
                                      on='name', how='inner')
    else:
        prTopic = product_topic.merge(prOther[['name', 'total', 'total_PSR', 'social', 'social_PSR', '电商', 'ec_PSR']],
                                      on='name', how='inner')
    prTopic = prTopic.sort_values(by='total', ascending=False)
    prTopic = prTopic[
        ['name', 'total', 'total_PSR', 'social', 'social_PSR', '电商', 'ec_PSR', 'Sizing', 'Design', 'Price', 'Quality',
         'Authenticity', 'Material', 'Flexibility', 'Fitting', 'Comfort',
         'Sizing_Pct', 'Design_Pct', 'Price_Pct', 'Quality_Pct', 'Authenticity_Pct', 'Material_Pct', 'Flexibility_Pct',
         'Fitting_Pct', 'Comfort_Pct']]
    prTopic = prTopic[~(prTopic['name'].isin(['Yeezy Boost', 'Superstar', 'UltraBOOST', 'Chuck 70', 'AlphaBounce',
                                              'ONE STAR', 'Stan Smith', 'Chuck Taylor', '悟道ACE', 'SK8']))]
    iloc_write(sheet10, prTopic, 3, 2)


def daily_volume(id_query, dayslist, workbook, wbFran, wbCam, writer):
    # queryName = pd.read_excel(r'R:\yuqing_fmcg\0_rpt_pkg\weekly_report\docs\query_id.xlsx')
    sheet1 = workbook.create_sheet('1TotalBuzz')
    sheet2 = workbook.create_sheet('2TopicBuzz')
    sheet2_1 = workbook.create_sheet('2TopicBuzzDetail')
    sheet3 = workbook.create_sheet('3ByDayBuzz')
    sheet4 = workbook.create_sheet('4ByDaySenti')

    sheet10 = wbFran.create_sheet("10FranKPI")
    sheet10_1 = wbFran.create_sheet("10FranKPIDetail")
    redFill = PatternFill(start_color='FFFF00', end_color = 'FFFF00', fill_type = 'solid')
    sheet10.conditional_formatting.add('E1:E50', FormulaRule(formula=['E1=Air Jordan 1'], fill=redFill))
    sheet11 = wbFran.create_sheet('11OthersBuzz')
    # sheet11_1 = wbFran.create_sheet('11OthersSenti')

    # cam6 = wbCam.create_sheet('6CamByDay')

    df = pd.DataFrame()
    pltDf = pd.DataFrame()
    senDf = pd.DataFrame()
    topicDf = pd.DataFrame()
    franTopic = pd.DataFrame()
    topicDetail = pd.DataFrame()
    franTopicDetail = pd.DataFrame()
    for id, queryList in id_query.items():
        type = id_type[str(id)]
        for query in queryList.split(','):
            # 获取到天的声量数据
            df = df.append(daily_buzz(headers, id, query, dayslist, type))
            # 获取每天的平台分布数据
            pltDf = pltDf.append(daily_plt(headers, id, query, dayslist, type))
            # 获取每天的情感分布
            senDf = senDf.append(sentiOpen(headers, id, query, dayslist, type, 'total'))
            senDf = senDf.append(sentiOpen(headers, id, query, dayslist, type, '5'))
            senDf = senDf.append(sentiOpen(headers, id, query, dayslist, type, '7,6,3,1,4,2'))
            # 获取首页的topic buzz
            if type == 'brand':
                topicDf = topicDf.append(topic_buzz(headers, id, query, dayslist, type, '7'))
                topicDetail = topicDetail.append(DetailTopic(headers, id, query, dayslist, '7'))
            if type == 'franchise':
                franTopic = franTopic.append(get_fran_topic(headers, id, query, dayslist, type, '7'))
                franTopicDetail = franTopicDetail.append(DetailTopicFran(headers, id, query, dayslist, '7'))


    # 输出query_name对应表
    queryName = df[['name', 'query']]
    queryName = queryName.drop_duplicates()
    queryName.to_excel(r'R:\yuqing_fmcg\3_report\queryName\query_names.xlsx', index=False)
    print('queryname已写出')

    pltDf = pltDf.merge(queryName, on='query', how='left')
    senDf = senDf.merge(queryName, on='query', how='left')
    topicDetail = topicDetail.merge(queryName, on='query', how='left')
    franTopicDetail = franTopicDetail.merge(queryName, on='query', how='left')

    pltDf.to_excel(writer, sheet_name='plt_daily', index=False)
    df.to_excel(writer, sheet_name='daily', index=False)
    senDf.to_excel(writer, sheet_name='sent_raw', index=False)
    topicDf.to_excel(writer, sheet_name='brand_topic', index=False)
    franTopic.to_excel(writer, sheet_name='fran_topic', index=False)
    topicDetail.to_excel(writer, sheet_name='detail_topic', index=False)
    franTopicDetail.to_excel(writer, sheet_name='fran_detail_topic', index=False)

    topicDetail = topicDetail.pivot_table(index=['query', 'name'], columns='1级属性', values = '3级属性总量', aggfunc='sum').reset_index()
    franTopicDetail = franTopicDetail.pivot_table(index=['query', 'name'], columns='1级属性', values = '2级属性总量', aggfunc='sum').reset_index()

    print(topicDetail.columns)
    # topicDetail = topicDetail.rename(columns={'Brand': 'Brand Reputation', 'Campaign': 'Campaign/Events', ''})
    topic_buzz_4_detail(topicDetail, sheet2_1)
    # nike, adi属性声量
    topic_buzz_4(topicDf, sheet2)
    sent = sent_from_daily(senDf, writer)


    # 产品kpi输出
    # product_topic = fran_topic(franTopic)
    product_topic = fran_topic_detail(franTopic)
    prDetailTopic = fran_topic_detail(franTopicDetail)
    # 到天的情感计算
    ncol = 2
    for brandName in ['Nike', 'Adidas', 'Lining', 'Anta']:
        daily_sent = senDf[(senDf['plt']=='social') & (senDf['name'] ==brandName)]
        daily_sent['PSR'] = list(map(lambda x, y: x / (x + y) if (x + y) > 0 else '-', daily_sent['正面'], daily_sent['负面']))
        for col in ['正面', '负面', '中性', 'PSR']:
            if col != 'PSR':
                daily_sent[col] = daily_sent[col].astype(int)
            daily_sent = daily_sent.rename(columns={col: '{}_{}'.format(brandName, col)})

        daily_sent = daily_sent.sort_values(by='date')

        if brandName == 'Nike':
            daily_sent = daily_sent[['date', f'{brandName}_正面', f'{brandName}_中性', f'{brandName}_负面', f'{brandName}_PSR']]
            iloc_write(sheet4, daily_sent, 2, ncol)
            ncol = ncol + 5
        else:
            daily_sent = daily_sent[
                [f'{brandName}_正面', f'{brandName}_中性', f'{brandName}_负面', f'{brandName}_PSR']]
            iloc_write(sheet4, daily_sent, 2, ncol)
            ncol = ncol + 4

    for typeStr in ['brand', 'franchise']:
        pltVol = pltDf[pltDf['type'] == typeStr].pivot_table(index='name', columns='platform', values='volume', aggfunc='sum').reset_index()
        pltVol['total'] = list(map(lambda a, b, c, d, e, f, g: int(a) + int(b) +int(c) + int(d) + int(e) + int(f) + int(g),
                                   pltVol['微信'],pltVol['微博'], pltVol['新闻'], pltVol['电商'], pltVol['视频'], pltVol['论坛'], pltVol['问答']))
        pltVol['social'] = list(map(lambda a, b, c, e, f, g: int(a) + int(b) +int(c) + int(e) + int(f) + int(g),
                                   pltVol['微信'],pltVol['微博'], pltVol['新闻'], pltVol['视频'], pltVol['论坛'], pltVol['问答']))
        pltVol['ec'] = pltVol['电商']

        if typeStr == 'brand':
            # 写出品牌top10声量, 如果originals, neo, jordan进top10，就延后
            # pltVol = pltVol.sort_values(by='total', ascending=False).iloc[0:13]
            # 定义输出表头顺序
            zhanbi_Acols = OrderedSet(['name', 'total', '电商', '微博', '微信', '新闻', '论坛', '视频', '问答', 'social',
       'ec', '电商占比', '微博占比', '微信占比', '新闻占比', '论坛占比', '视频占比', '问答占比'])
            for col in list(set(pltVol.columns)-set(['name', 'total', 'social', 'ec'])):
                pltVol['{}占比'.format(col)] = list(map(lambda x, y: float(x/y), pltVol[col], pltVol['total']))
            zhanbi_cols = list(zhanbi_Acols & set(pltVol.columns))
            pltVol = pltVol[zhanbi_cols]
            iloc_write(sheet1, pltVol, 3, 2)
            # 写出品牌top10情感
            bdsent = sent[sent['name'].isin(pltVol['name'].unique())].drop('type', 1)
            iloc_write(sheet1, bdsent, 58, 2)
        # 输出产品声量
        else:
            prOther = pltVol.merge(sent[['name', 'total_PSR', 'social_PSR', 'ec_PSR']], on='name', how='inner')
            prOther = prOther[['name', '微信', '微博', '论坛', '新闻', '视频', '问答', 'social', 'social_PSR', '电商', 'ec_PSR', 'total', 'total_PSR']]
            prOther = prOther.sort_values(by='total', ascending=False)
            iloc_write(sheet11, prOther, 3, 2)

            sheet10_writer(product_topic, prOther, sheet10)
            sheet10_writer(prDetailTopic, prOther, sheet10_1)
            # prTopic = product_topic.merge(prOther[['name', 'total', 'total_PSR', 'social', 'social_PSR', '电商', 'ec_PSR']], on='name', how='inner')
            # prTopic = prTopic.sort_values(by='total', ascending=False)
            # prTopic = prTopic[['name', 'total', 'total_PSR', 'social', 'social_PSR', '电商', 'ec_PSR', 'Sizing', 'Design', 'Price', 'Quality', 'Authenticity', 'Material', 'Flexibility', 'Fitting', 'Comfort',
            #                    'Sizing_Pct', 'Design_Pct', 'Price_Pct', 'Quality_Pct', 'Authenticity_Pct', 'Material_Pct', 'Flexibility_Pct', 'Fitting_Pct', 'Comfort_Pct']]
            # prTopic = prTopic[~(prTopic['name'].isin(['Yeezy Boost', 'Superstar', 'UltraBOOST', 'Chuck 70', 'AlphaBounce',
            #                                         'ONE STAR', 'Stan Smith', 'Chuck Taylor', '悟道ACE', 'SK8']))]
            # iloc_write(sheet10, prTopic, 3, 2)
            #
            # prsent = sent[sent['name'].isin(pltVol['name'].unique())].drop('type', 1)
            # iloc_write(sheet11_1, prsent, 3, 2)

    daily_volume_from_plt(pltDf, sheet3)

def sentByTopic(id_query, dayslist, writer, wbFran):
    sheet12 = wbFran.create_sheet("12FranTopicSenti")
    df = pd.DataFrame()
    topicList = ['Sizing', 'Design', 'Price', 'Quality', 'Authenticity', 'Material', 'Flexibility', 'Comfort']
    queryName = pd.read_excel(r'R:\yuqing_fmcg\0_rpt_pkg\weekly_report\docs\query_names.xlsx')
    for id, queryList in id_query.items():
        type = id_type[str(id)]
        for query in queryList.split(','):
            df = df.append(sentiOpenTopic(headers, id, query, dayslist, type, topicList))
    df['query'] = df['query'].astype(str)
    queryName['query'] = queryName['query'].astype(str)
    df = df.merge(queryName, on='query', how='left')
    df.to_excel(writer, sheet_name='franTopicSenti', index=False)

    result = df.pivot_table(index=['name', 'topic'], values=['正面', '中性', '负面'], aggfunc='sum').reset_index()
    result['PSR'] = list(map(lambda x, y: float(x / (x + y)) if (x + y) > 0 else '-', result['正面'], result['负面']))
    for col in ['正面', '负面', '中性', 'PSR']:
        if col != 'PSR':
            result[col] = result[col].astype(int)
    result = result[result['name'].isin(['Air Force1', 'Air Jordan 1', 'Air Jordan Retro', 'Air Max Others', 'Kyrie', 'Air Jordan 11'])]
    iloc_write(sheet12, result, 3, 2)




if __name__ =='__main__':
    timeList = pd.date_range('2019-12-01', '2019-12-31')
    daysList =[i.strftime("%Y-%m-%d") for i in timeList]
    # print(daysList)
    print(daysList)
    outPath = os.path.join(r'R:\yuqing_fmcg\3_report\2_Monthly_report', daysList[-1].replace("-", '')[:6], 'orig')
    if not os.path.exists(outPath):
        os.makedirs(outPath)
    wbAll = Workbook()
    wbFran = Workbook()
    wbCam = Workbook()
    writer = pd.ExcelWriter(os.path.join(outPath, 'Month_rpt_ref_{}_WB.xlsx'.format(daysList[-1])))
    daily_volume(id_query, daysList, wbAll, wbFran, wbCam, writer)
    sentByTopic(id_query, daysList, writer, wbFran)
    wbAll.save(os.path.join(outPath, 'Nike_Mon_overall_WB.xlsx'))
    wbFran.save(os.path.join(outPath,'Nike_Mon_Fran.xlsx_WB'))
    wbCam.save(os.path.join(outPath, 'Nike_Mon_Campaign.xlsx_WB'))
    writer.save()

